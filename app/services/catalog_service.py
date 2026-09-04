"""Сервис каталога: парсинг цен поставщика, работа с прайс-листами.

Используется роутами /prices, скриптом scripts/import_catalog.py и каталогом.
"""

import openpyxl
from decimal import Decimal, InvalidOperation
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import PriceList, Product, ProductCategory, ProductPrice

# Колонки xlsx-каталога поставщика (формат фазы 17)
COL_CATEGORY, COL_NAME, COL_SKU, COL_PRICE, COL_URL, COL_IMAGE = 1, 4, 5, 6, 7, 8


def parse_price_amount(raw) -> Decimal | None:
    """«1 234,56» / «1234.56» / «уточняйте» → Decimal | None.

    Поставщики пишут цены кто во что горазд: пробелы-разделители, запятая
    как десятичный разделитель, текст-заглушки. Не число — значит цены нет.
    """
    if raw is None:
        return None
    s = str(raw).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        value = Decimal(s)
    except InvalidOperation:
        return None
    return value if value > 0 else None


def read_catalog_rows(path: str | Path, limit: int = 0) -> list[dict]:
    """Читает xlsx-каталог поставщика (формат фазы 17) в список dict."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[COL_NAME] is None or r[COL_URL] is None:
            continue
        rows.append({
            "category": str(r[COL_CATEGORY]).strip() if r[COL_CATEGORY] else None,
            "name": str(r[COL_NAME]).strip(),
            "sku": str(r[COL_SKU]).strip() if r[COL_SKU] else None,
            "price_raw": r[COL_PRICE],
            "url": str(r[COL_URL]).strip(),
            "image": str(r[COL_IMAGE]).strip() if r[COL_IMAGE] else None,
        })
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows


async def get_or_create_default_pricelist(session: AsyncSession) -> PriceList:
    """Дефолтный прайс-лист — ровно один (гонку не рассматриваем: создание
    идёт под админскими действиями, is_default=True уникален практически)."""
    pricelist = (
        await session.execute(
            select(PriceList).where(PriceList.is_default == True)  # noqa: E712
        )
    ).scalar_one_or_none()
    if pricelist is None:
        pricelist = PriceList(name="Базовый", currency="RUB", is_default=True)
        session.add(pricelist)
        await session.flush()
    return pricelist


async def upsert_price(session: AsyncSession, product_id: int, price_list_id: int,
                       value: Decimal | None = None, price_in: Decimal | None = None,
                       price_out: Decimal | None = None, vat_rate: Decimal | None = None) -> str:
    """Возвращает 'created' | 'updated'. Раскладка (фаза 22): price_in — входящая
    б/НДС, price_out — отпускная б/НДС, vat_rate — %НДС; price (отпускная С НДС,
    её берут КП/счёт) пересчитывается = price_out * (1 + vat/100). Прямое value —
    fallback для совместимости со старым импортом."""
    pp = (
        await session.execute(
            select(ProductPrice).where(
                ProductPrice.product_id == product_id,
                ProductPrice.price_list_id == price_list_id,
            )
        )
    ).scalar_one_or_none()
    if pp is None:
        pp = ProductPrice(product_id=product_id, price_list_id=price_list_id, price=value or 0)
        session.add(pp)
        status = "created"
    else:
        status = "updated"
    if price_in is not None:
        pp.price_in = price_in
    if price_out is not None:
        pp.price_out = price_out
    if vat_rate is not None:
        pp.vat_rate = vat_rate
    if pp.price_out is not None:
        pp.price = (Decimal(str(pp.price_out)) * (Decimal("100") + Decimal(str(pp.vat_rate or 22))) / Decimal("100")).quantize(Decimal("0.01"), rounding="ROUND_HALF_UP")
    elif value is not None:
        pp.price = value
    return status


async def find_product(session: AsyncSession, source_url: str | None, sku: str | None) -> Product | None:
    """Товар по source_url (надёжно), затем по первому артикулу (артикулы у
    поставщика неуникальны — берём первый активный)."""
    if source_url:
        product = (
            await session.execute(select(Product).where(Product.source_url == source_url))
        ).scalar_one_or_none()
        if product:
            return product
    if sku:
        product = (
            await session.execute(
                select(Product).where(Product.sku == sku, Product.is_active == True)  # noqa: E712
                .order_by(Product.id).limit(1)
            )
        ).scalar_one_or_none()
        if product:
            return product
    return None


async def upsert_category(session: AsyncSession, name: str, sort_order: int) -> ProductCategory:
    cat = (
        await session.execute(
            select(ProductCategory).where(
                ProductCategory.name == name, ProductCategory.parent_id.is_(None)
            )
        )
    ).scalar_one_or_none()
    if cat is None:
        cat = ProductCategory(name=name, parent_id=None, sort_order=sort_order)
        session.add(cat)
        await session.flush()
    return cat


async def upsert_product(session: AsyncSession, cat: ProductCategory | None, row: dict, brand: str) -> str:
    """Возвращает 'created' | 'updated'. Каталожный upsert матчит ТОЛЬКО по
    source_url: артикулы поставщика неуникальны (в каталоге АгроВиты 16 строк
    делят артикул с другой строкой) — sku-фолбэк здесь сливал разные товары.
    Фолбэк по sku остаётся только в find_product (импорт цен)."""
    product = None
    if row.get("url"):
        product = (
            await session.execute(select(Product).where(Product.source_url == row["url"]))
        ).scalar_one_or_none()
    if product is None:
        product = Product(
            name=row["name"], source_url=row.get("url"), brand=brand,
            category_id=cat.id if cat else None,
        )
        session.add(product)
        status = "created"
    else:
        status = "updated"
        product.name = row["name"]
        if cat:
            product.category_id = cat.id
    product.sku = row.get("sku") or None
    return status

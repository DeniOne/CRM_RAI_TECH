"""Импорт каталога производителя из xlsx в БД CRM (фаза 17).

Формат файла (лист любой, колонки по индексам):
  0: № | 1: Категория | 2: Подкатегория | 3: Полный путь | 4: Название |
  5: Артикул | 6: Цена (игнорируется, цены — фаза 18) | 7: Ссылка | 8: Картинка

Идемпотентность: upsert по source_url (в каталоге АгроВиты уникален — проверено).
Картинки качаются в storage/catalog/images/ с троттлингом, ретраями и
возобновляемостью (существующие файлы не перекачиваются).

Примеры:
  python scripts/import_catalog.py --xlsx path.xlsx --limit 25
  python scripts/import_catalog.py --xlsx path.xlsx --no-images
  python scripts/import_catalog.py --images-only
"""

import argparse
import asyncio
import json
import random
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlsplit

import httpx
import openpyxl
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import settings  # noqa: E402
from app.database import async_session_maker, init_db  # noqa: E402
from app.models import Product, ProductCategory  # noqa: E402

DEFAULT_XLSX = r"I:\Мой диск\РАИ Технологии\РАИ Технологии\АгроВитаСервис\agrovita_catalog.xlsx"

COL_CATEGORY, COL_NAME, COL_SKU, COL_URL, COL_IMAGE = 1, 4, 5, 7, 8

# Каталог производителя — качаем вежливо: браузерный UA (часть сайтов режет
# нестандартные UA на уровне WAF), пауза между запросами, ретраи с backoff.
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "image/avif,image/webp,image/png,image/jpeg,*/*;q=0.8",
}
IMG_EXTS = {"jpg", "jpeg", "png", "webp"}
IMAGE_MAGIC = (
    (b"\xff\xd8\xff", "jpg"),
    (b"\x89PNG", "png"),
    (b"GIF8", "gif"),
)


def read_catalog_rows(path: str, limit: int = 0) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[COL_NAME] is None or r[COL_URL] is None:
            continue
        rows.append({
            "category": str(r[COL_CATEGORY]).strip() if r[COL_CATEGORY] else None,
            "name": str(r[COL_NAME]).strip(),
            "sku": str(r[COL_SKU]).strip() if r[COL_SKU] else None,
            "url": str(r[COL_URL]).strip(),
            "image": str(r[COL_IMAGE]).strip() if r[COL_IMAGE] else None,
        })
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows


def image_ext(url: str) -> str:
    suffix = urlsplit(url).path.rsplit(".", 1)[-1].lower()
    return suffix if suffix in IMG_EXTS else "jpg"


def sniff_ext(data: bytes) -> str | None:
    for magic, ext in IMAGE_MAGIC:
        if data.startswith(magic):
            return ext
    if len(data) > 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "webp"
    return None


async def upsert_category(session, name: str, sort_order: int) -> ProductCategory:
    cat = (
        await session.execute(
            select(ProductCategory).where(
                ProductCategory.name == name, ProductCategory.parent_id.is_(None)
            )
        )
    ).scalar_one_or_none()
    if cat is None:
        cat = ProductCategory(name=name, parent_id=None, sort_order=sort_order)
        session.add(cat)
        await session.flush()
    return cat


async def upsert_product(session, cat: ProductCategory | None, row: dict, brand: str) -> str:
    """Возвращает 'created' | 'updated'."""
    product = None
    if row["url"]:
        product = (
            await session.execute(select(Product).where(Product.source_url == row["url"]))
        ).scalar_one_or_none()
    if product is None:
        product = Product(
            name=row["name"], source_url=row["url"], brand=brand,
            category_id=cat.id if cat else None,
        )
        session.add(product)
        status = "created"
    else:
        status = "updated"
        product.name = row["name"]
        if cat:
            product.category_id = cat.id
    product.sku = row["sku"] or None
    return status


async def download_images(session, products: list[Product], delay: float, report: dict) -> None:
    """Последовательно, с паузой delay + джиттер: 1300+ файлов без троттлинга —
    прямой путь к 429/бану по IP. Возобновляемо: существующие файлы пропускаются,
    повторный запуск докачивает остаток."""
    images_dir = settings.CATALOG_IMAGES_DIR
    images_dir.mkdir(parents=True, exist_ok=True)

    # source_url_image — transient-атрибут (URL из xlsx; в БД хранится только
    # локальное имя файла). Без него товар в очередь загрузки не попадает.
    todo = [p for p in products if getattr(p, "source_url_image", None) and not p.image_file]
    already = sum(1 for p in products if p.image_file)
    report["images_skipped_existing"] = already
    print(f"К загрузке: {len(todo)} картинок (уже есть: {already})")

    timeout = httpx.Timeout(connect=10, read=20, write=20, pool=10)
    ok = failed = 0
    started = time.monotonic()

    async with httpx.AsyncClient(headers=HEADERS, timeout=timeout, follow_redirects=True) as client:
        for i, p in enumerate(todo, 1):
            data = None
            for attempt in range(3):
                try:
                    resp = await client.get(p.source_url_image)
                    if resp.status_code == 200 and resp.content:
                        data = resp.content
                        break
                    if resp.status_code == 404:
                        break  # 404 ретраить бессмысленно
                except httpx.HTTPError:
                    pass
                await asyncio.sleep(min(2 ** attempt, 7))
            await asyncio.sleep(delay + random.uniform(0, delay * 0.5))

            if data:
                ext = sniff_ext(data) or image_ext(p.source_url_image)
                fname = f"{p.id}.{ext}"
                (images_dir / fname).write_bytes(data)
                p.image_file = fname
                ok += 1
            else:
                failed += 1
                report["failed_images"].append({"product_id": p.id, "url": p.source_url_image})

            if i % 50 == 0:
                await session.commit()
                rate = i / max(time.monotonic() - started, 1)
                eta = (len(todo) - i) / max(rate, 0.01) / 60
                print(f"  {i}/{len(todo)} (ok={ok}, fail={failed}), ~{eta:.0f} мин до конца")

    await session.commit()
    report["images_ok"], report["images_failed"] = ok, failed


async def main() -> None:
    parser = argparse.ArgumentParser(description="Импорт каталога товаров из xlsx")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="путь к xlsx-каталогу")
    parser.add_argument("--brand", default="АгроВита", help="бренд/поставщик")
    parser.add_argument("--limit", type=int, default=0, help="обработать только N строк (тест)")
    parser.add_argument("--no-images", action="store_true", help="не качать картинки")
    parser.add_argument("--images-only", action="store_true", help="только докачать картинки")
    parser.add_argument("--delay", type=float, default=0.6, help="пауза между запросами картинок, сек")
    args = parser.parse_args()

    started = time.monotonic()
    await init_db()

    report = {
        "xlsx": args.xlsx, "brand": args.brand, "started": datetime.now().isoformat(),
        "rows": 0, "created": 0, "updated": 0, "categories": 0,
        "images_ok": 0, "images_failed": 0, "images_skipped_existing": 0,
        "failed_images": [],
    }

    async with async_session_maker() as session:
        if args.images_only:
            # Только докачка: соответствие товар↔URL картинки берём из xlsx
            url_to_image = {
                r["url"]: r["image"] for r in read_catalog_rows(args.xlsx) if r["image"]
            }
            products = (
                (await session.execute(select(Product).where(Product.source_url.is_not(None))))
                .scalars()
                .all()
            )
            for p in products:
                p.source_url_image = url_to_image.get(p.source_url)
        else:
            rows = read_catalog_rows(args.xlsx, args.limit)
            cats_order: dict[str, int] = {}
            cat_cache: dict[str, ProductCategory] = {}
            for row in rows:
                if row["category"] and row["category"] not in cat_cache:
                    name = row["category"]
                    if name not in cats_order:
                        cats_order[name] = len(cats_order)
                    cat_cache[name] = await upsert_category(session, name, cats_order[name])
                status = await upsert_product(session, cat_cache.get(row["category"]), row, args.brand)
                report["rows"] += 1
                report[status] += 1
            report["categories"] = len(cat_cache)
            await session.commit()
            print(f"Товаров обработано: {report['rows']} (новых {report['created']}, обновлено {report['updated']})")
            print(f"Категорий: {report['categories']}")

        if not args.no_images:
            if not args.images_only:
                url_to_image = {r["url"]: r["image"] for r in rows if r["image"]}
                products = (
                    (await session.execute(select(Product).where(Product.source_url.is_not(None))))
                    .scalars()
                    .all()
                )
                for p in products:
                    p.source_url_image = url_to_image.get(p.source_url)
            await download_images(session, products, args.delay, report)

    report["elapsed_sec"] = round(time.monotonic() - started, 1)

    settings.STORAGE_DIR.joinpath("exports").mkdir(parents=True, exist_ok=True)
    report_path = (
        settings.STORAGE_DIR / "exports" / f"catalog_import_{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    )
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Картинок: ok={report['images_ok']}, fail={report['images_failed']}, "
          f"уже были={report['images_skipped_existing']}")
    if report["failed_images"]:
        print(f"Битые URL ({len(report['failed_images'])}) — первые 5:")
        for f in report["failed_images"][:5]:
            print(f"  product {f['product_id']}: {f['url']}")
    print(f"Отчёт: {report_path}")
    print(f"Готово за {report['elapsed_sec']} c")


if __name__ == "__main__":
    asyncio.run(main())
